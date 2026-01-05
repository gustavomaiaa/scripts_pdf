import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def exportar_excel(lista_dados, caminho_saida):
    # Para criar dataframe

    df = pd.DataFrame(lista_dados)

    # ordem das colunas
    colunas = [
        "Municipio",
        "Estado",
        "Data",
        "Orcamento",
        "Prefeito",
        "Secretario_Financeiro",
        "Arquivo"
    ]

    df = df[colunas]

    # exportar p/ excel
    df.to_excel(caminho_saida, index=False, sheet_name="Relatório")

    # formatação c/ openpyxl
    wb = load_workbook(caminho_saida)
    ws = wb.active

    # congelar cabeçalho
    ws.freeze_panes = "A2"

    # filtro automático
    ws.auto_filter.ref = ws.dimensions

    # estilo do cabeçalho
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # autoajuste das colunas
    for col in ws.columns:
        max_lenght = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_lenght = max(max_lenght, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_lenght + 3

    # formtação da coluna orçamento (R$)
    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].number_format = '[$R$-pt-BR] #,##0.00'

    ws.column_dimensions["D"].width = 18

    # formatação da coluna data
    for row in range(2, ws.max_row + 1):
        ws[f"C{row}"].number_format = 'DD/MM/YYYY'

    # salvando...
    wb.save(caminho_saida)